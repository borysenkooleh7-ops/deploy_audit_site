"""
Servicio para importar marcas de auditoría desde archivos Excel con filtrado basado en colores.

Reglas de Color:
- Filas VERDES: Importar (marcas activas)
- Filas BLANCAS/Sin color: Omitir (marcas inactivas/borradores)
- Filas AMARILLAS: Omitir (ejemplos/plantillas)
"""

from openpyxl import load_workbook
from django.db import transaction
from django.core.exceptions import ValidationError
from auditoria.models import AuditMark
from django.db.models import Q
import logging

logger = logging.getLogger(__name__)


class AuditMarkImportService:
    """
    Servicio para importar marcas de auditoría desde Excel con filtrado de color.
    """

    # Umbrales de detección de color
    GREEN_COLORS = [
        '00FF00', 'C6EFCE', '90EE90', '92D050', 'C5E0B4',
        '00B050', 'E2EFDA', 'A9D08E', '70AD47'
    ]
    YELLOW_COLORS = [
        'FFFF00', 'FFFFE0', 'FFF4CE', 'FFEB9C', 'FFD966',
        'FFC000', 'F4B084', 'FCE4D6'
    ]

    def __init__(self, audit_id, excel_file):
        """
        Inicializar el servicio de importación.

        Args:
            audit_id: ID de la auditoría para vincular las marcas
            excel_file: Objeto de archivo Excel cargado
        """
        self.audit_id = audit_id
        self.excel_file = excel_file
        self.marks_imported = 0
        self.marks_skipped_white = 0
        self.marks_skipped_yellow = 0
        self.marks_skipped_invalid = 0
        self.errors = []

    def validate_file(self):
        """
        Validar el archivo Excel cargado.

        Raises:
            ValidationError: Si el archivo no es válido
        """
        # Verificar extensión del archivo
        filename = self.excel_file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            raise ValidationError("El archivo debe ser formato Excel (.xlsx o .xls)")

        # Verificar tamaño del archivo (máx 5MB)
        if self.excel_file.size > 5 * 1024 * 1024:
            raise ValidationError("El archivo no debe superar 5MB")

        # Intentar cargar el libro de trabajo para verificar que no está corrupto
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            wb.close()
        except Exception as e:
            raise ValidationError(f"El archivo está corrupto o no se puede leer: {str(e)}")

    def _get_cell_background_color(self, cell):
        """
        Extraer el color de fondo de la celda.

        Args:
            cell: Objeto de celda de openpyxl

        Returns:
            str: Código de color hexadecimal sin '#' o None
        """
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color

            # Manejar colores RGB
            if hasattr(color, 'rgb') and color.rgb:
                rgb = str(color.rgb)
                # Eliminar canal alfa si está presente (formato ARGB)
                if len(rgb) == 8:
                    return rgb[2:].upper()  # Eliminar los primeros 2 caracteres (alfa)
                return rgb.upper()

        return None

    def _is_green_row(self, row_cells):
        """
        Verificar si la fila tiene fondo verde (debe ser importada).

        Args:
            row_cells: Lista de objetos de celda de openpyxl

        Returns:
            bool: True si alguna celda de la fila tiene fondo verde
        """
        for cell in row_cells:
            bg_color = self._get_cell_background_color(cell)
            if bg_color and any(green in bg_color for green in self.GREEN_COLORS):
                return True
        return False

    def _is_yellow_row(self, row_cells):
        """
        Verificar si la fila tiene fondo amarillo (ejemplo/plantilla, debe omitirse).

        Args:
            row_cells: Lista de objetos de celda de openpyxl

        Returns:
            bool: True si alguna celda de la fila tiene fondo amarillo
        """
        for cell in row_cells:
            bg_color = self._get_cell_background_color(cell)
            if bg_color and any(yellow in bg_color for yellow in self.YELLOW_COLORS):
                return True
        return False

    def _is_white_or_no_color_row(self, row_cells):
        """
        Verificar si la fila tiene fondo blanco o sin color (debe omitirse).

        Args:
            row_cells: Lista de objetos de celda de openpyxl

        Returns:
            bool: True si la fila no tiene color o tiene fondo blanco
        """
        # Si no es verde y no es amarilla, se considera blanca/sin color
        return not self._is_green_row(row_cells) and not self._is_yellow_row(row_cells)

    def parse_excel(self):
        """
        Analizar el archivo Excel y extraer las marcas de auditoría basándose en los colores de las filas.

        Returns:
            list: Lista de diccionarios de marcas para importar
        """
        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb.active

        marks = []

        # Iterar a través de las filas (omitir fila de encabezado 1)
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                # Verificar color de la fila
                if self._is_yellow_row(row_cells):
                    logger.debug(f"Fila {row_idx}: AMARILLO - Omitiendo (ejemplo)")
                    self.marks_skipped_yellow += 1
                    continue

                if self._is_white_or_no_color_row(row_cells):
                    # Verificar si la fila tiene algún dato
                    has_data = any(cell.value for cell in row_cells[:4])
                    if has_data:
                        logger.debug(f"Fila {row_idx}: BLANCO - Omitiendo (inactivo)")
                        self.marks_skipped_white += 1
                    continue

                if self._is_green_row(row_cells):
                    logger.debug(f"Fila {row_idx}: VERDE - Procesando")
                    # Extraer datos de marca de la fila
                    mark_data = self._extract_mark_from_row(row_cells, row_idx)
                    if mark_data:
                        marks.append(mark_data)
                    else:
                        self.marks_skipped_invalid += 1

            except Exception as e:
                error_msg = f"Fila {row_idx}: Error al procesar - {str(e)}"
                logger.error(error_msg)
                self.errors.append(error_msg)
                self.marks_skipped_invalid += 1

        wb.close()
        return marks

    def _extract_mark_from_row(self, row_cells, row_idx):
        """
        Extraer datos de marca de una sola fila.

        Args:
            row_cells: Lista de objetos de celda de openpyxl
            row_idx: Índice de fila para informe de errores

        Returns:
            dict: Datos de marca o None si no es válido
        """
        # Extraer valores (manejar valores None)
        symbol = str(row_cells[0].value).strip() if row_cells[0].value else None
        description = str(row_cells[1].value).strip() if row_cells[1].value else None
        work_paper = str(row_cells[2].value).strip() if len(row_cells) > 2 and row_cells[2].value else None
        category = str(row_cells[3].value).strip() if len(row_cells) > 3 and row_cells[3].value else None

        # Validar campos requeridos
        if not symbol or not description:
            logger.warning(f"Fila {row_idx}: Faltan campos requeridos (símbolo o descripción)")
            return None

        # Validación adicional: Omitir si todavía contiene "Ejemplo:" incluso en fila verde
        if "Ejemplo:" in description or "Example:" in description:
            logger.warning(f"Fila {row_idx}: Contiene 'Ejemplo:' - Omitiendo")
            return None

        return {
            'symbol': symbol,
            'description': description,
            'work_paper_number': work_paper,
            'category': category,
            'row_number': row_idx  # Para depuración
        }

    @transaction.atomic
    def import_marks(self, replace_existing=True):
        """
        Importar marcas a la base de datos con seguridad de transacción.

        Args:
            replace_existing: Si es True, eliminar las marcas existentes para esta auditoría

        Returns:
            dict: Resultados de importación con estadísticas
        """
        # Validar archivo primero
        self.validate_file()

        # Analizar Excel para extraer marcas
        marks = self.parse_excel()

        if replace_existing:
            # Eliminar marcas existentes para esta auditoría
            deleted_count = AuditMark.objects.filter(audit_id=self.audit_id).delete()[0]
            logger.info(f"Eliminadas {deleted_count} marcas existentes para auditoría {self.audit_id}")

        # Creación masiva de nuevas marcas
        if marks:
            mark_objects = [
                AuditMark(
                    audit_id=self.audit_id,
                    symbol=mark['symbol'],
                    description=mark['description'],
                    work_paper_number=mark['work_paper_number'],
                    category=mark['category'],
                    is_active=True
                )
                for mark in marks
            ]

            AuditMark.objects.bulk_create(mark_objects)
            self.marks_imported = len(mark_objects)
            logger.info(f"Importadas {self.marks_imported} marcas para auditoría {self.audit_id}")

        return {
            'success': True,
            'marks_imported': self.marks_imported,
            'marks_skipped_white': self.marks_skipped_white,
            'marks_skipped_yellow': self.marks_skipped_yellow,
            'marks_skipped_invalid': self.marks_skipped_invalid,
            'errors': self.errors
        }

    def get_summary_message(self):
        """
        Obtener resumen legible de la operación de importación.

        Returns:
            str: Mensaje de resumen
        """
        return (
            f"✅ Importadas: {self.marks_imported} marcas\n"
            f"⚪ Omitidas (blancas/sin color): {self.marks_skipped_white}\n"
            f"🟡 Omitidas (ejemplos amarillos): {self.marks_skipped_yellow}\n"
            f"❌ Omitidas (inválidas): {self.marks_skipped_invalid}"
        )
