"""
Servicio para generar plantilla Excel para marcas de auditoría con codificación de colores.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.http import HttpResponse
import io


class AuditMarkTemplateGenerator:
    """
    Genera plantilla Excel para cargar marcas de auditoría.
    """

    # Definiciones de color
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")
    WHITE_FILL = PatternFill(fill_type=None)  # Sin relleno
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    BOLD_FONT = Font(bold=True)
    NORMAL_FONT = Font(size=11)

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Marcas de Auditoría"

    def _create_header(self):
        """Crear fila de encabezado con nombres de columna."""
        headers = ["Símbolo", "Descripción", "Papel de Trabajo", "Categoría"]

        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # Establecer anchos de columna
        self.ws.column_dimensions['A'].width = 12  # Símbolo
        self.ws.column_dimensions['B'].width = 50  # Descripción
        self.ws.column_dimensions['C'].width = 18  # Papel de Trabajo
        self.ws.column_dimensions['D'].width = 15  # Categoría

    def _add_example_rows(self):
        """Agregar filas de ejemplo con diferentes colores."""

        # FILAS VERDES (se importarán)
        green_examples = [
            ("✓", "Verificado con extracto bancario", "1", "Efectivo"),
            ("◊", "Revisado y confirmado", "2", "Efectivo"),
            ("★", "Referenciado cruzado con libro mayor", "10", "Egresos"),
            ("⊕", "Verificado matemáticamente", "11", "Egresos"),
        ]

        row = 2
        for symbol, desc, wp, cat in green_examples:
            self._add_row(row, symbol, desc, wp, cat, self.GREEN_FILL)
            row += 1

        # FILA BLANCA (se omitirá - inactiva)
        self._add_row(row, "×", "Marca inactiva - no se importará", "C-1", "Cuentas por cobrar", self.WHITE_FILL)
        row += 1

        # Más FILAS VERDES
        self._add_row(row, "▲", "Examinado con documentación de respaldo", "D-1", "Propiedad", self.GREEN_FILL)
        row += 1

        # FILAS AMARILLAS (ejemplos - se omitirán)
        yellow_examples = [
            ("❌", "Ejemplo: Marca de ejemplo 1", "TEST-1", "Ejemplo"),
            ("⚠", "Ejemplo: Marca de ejemplo 2", "TEST-2", "Ejemplo"),
        ]

        for symbol, desc, wp, cat in yellow_examples:
            self._add_row(row, symbol, desc, wp, cat, self.YELLOW_FILL)
            row += 1

    def _add_row(self, row_num, symbol, description, work_paper, category, fill):
        """Agregar una sola fila de datos con estilo."""
        data = [symbol, description, work_paper, category]

        for col_idx, value in enumerate(data, start=1):
            cell = self.ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.fill = fill
            cell.font = self.NORMAL_FONT
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = self.BORDER

    def _add_instructions_sheet(self):
        """Agregar una segunda hoja con instrucciones."""
        instructions_ws = self.wb.create_sheet("Instrucciones")
        instructions_ws.column_dimensions['A'].width = 80

        instructions = [
            ("INSTRUCCIONES PARA USAR ESTA PLANTILLA", self.HEADER_FONT),
            ("", None),
            ("1. COLOR DE FILAS:", self.BOLD_FONT),
            ("   🟢 VERDE: Marque las filas que desea importar al sistema", None),
            ("   ⚪ BLANCO/SIN COLOR: Estas filas NO se importarán (marcas inactivas)", None),
            ("   🟡 AMARILLO: Estas filas NO se importarán (son ejemplos de referencia)", None),
            ("", None),
            ("2. COLUMNAS REQUERIDAS:", self.BOLD_FONT),
            ("   • Símbolo: Carácter o emoji que representa la marca (ej: ✓, ◊, ★)", None),
            ("   • Descripción: Texto que explica qué significa esta marca", None),
            ("   • Papel de Trabajo: Código del papel (ej: 1, 10, A-1) - usado para emparejar con documentos", None),
            ("   • Categoría: Agrupación opcional (ej: Efectivo, Inventario)", None),
            ("", None),
            ("3. CÓMO MARCAR FILAS COMO VERDES:", self.BOLD_FONT),
            ("   a. Seleccione toda la fila", None),
            ("   b. Clic derecho → Formato de celdas → Relleno", None),
            ("   c. Seleccione color verde claro (ej: #C6EFCE)", None),
            ("", None),
            ("4. EJEMPLOS:", self.BOLD_FONT),
            ("   • Vea la pestaña 'Marcas de Auditoría' para ejemplos con colores", None),
            ("   • Las filas verdes en la plantilla serán importadas", None),
            ("   • Las filas amarillas son ejemplos y serán ignoradas", None),
            ("", None),
            ("5. DESPUÉS DE COMPLETAR:", self.BOLD_FONT),
            ("   • Guarde el archivo", None),
            ("   • En el sistema, vaya a su auditoría", None),
            ("   • Clic en 'Marcas de Auditoría'", None),
            ("   • Suba este archivo Excel", None),
            ("", None),
            ("6. NÚMEROS DE PAPEL DE TRABAJO:", self.BOLD_FONT),
            ("   • Use el número/nombre del archivo sin extensión", None),
            ("   • Ejemplo: Para archivo '1 PROGRAMA.docx' use '1'", None),
            ("   • Ejemplo: Para archivo '10 INTEGRACION EGRESOS.xlsx' use '10'", None),
            ("   • El sistema normalizará automáticamente (quita espacios, guiones, etc.)", None),
        ]

        for row_idx, (text, font) in enumerate(instructions, start=1):
            cell = instructions_ws.cell(row=row_idx, column=1)
            cell.value = text
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def generate(self):
        """
        Generar la plantilla completa.

        Returns:
            Workbook: Objeto de libro de trabajo openpyxl
        """
        self._create_header()
        self._add_example_rows()
        self._add_instructions_sheet()

        # Establecer hoja activa de vuelta a la hoja principal
        self.wb.active = 0

        return self.wb

    def get_http_response(self, filename="plantilla_marcas_auditoria.xlsx"):
        """
        Generar plantilla y devolver como respuesta HTTP para descarga.

        Args:
            filename: Nombre del archivo para descargar

        Returns:
            HttpResponse: Respuesta HTTP de Django con archivo Excel
        """
        wb = self.generate()

        # Guardar en buffer BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Crear respuesta
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
